#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Métodos auxiliares"""

from datetime import datetime
import os
import json
import re

from openpyxl import load_workbook
from six.moves.urllib_parse import urlparse

from six import string_types
from unidecode import unidecode

ABSOLUTE_PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
ABSOLUTE_SCHEMA_DIR = os.path.join(ABSOLUTE_PROJECT_DIR, "schemas")
STOP_WORDS = [
    "el", "la", "los", "las", "de", "del", "y", "a", "un", "una", "en"
]


def title_to_name(title, decode=True):
    """Convierte un título en un nombre normalizado para generar urls."""
    # decodifica y pasa a minúsculas
    if decode:
        title = unidecode(title)
    title = title.lower()

    # remueve caracteres no permitidos
    filtered_title = re.sub(r'[^a-z0-9- ]+', '', title)

    # remueve stop words y espacios y une palabras sólo con un "-"
    normalized_title = '-'.join(
        [word for word in filtered_title.split() if word not in STOP_WORDS])

    return normalized_title


def parse_date_string(date_string):
    """Parsea un string de una fecha con el formato de la norma
    ISO 8601 (es decir, las fechas utilizadas en los catálogos) en un
    objeto datetime de la librería estándar de python. Se tiene en cuenta
    únicamente la fecha y se ignora completamente la hora.

    Args:
        date_string (str): fecha con formato ISO 8601.

    Returns:
        datetime: objeto fecha especificada por date_string.
    """

    if not date_string:
        return None

    # La fecha cumple con la norma ISO 8601: YYYY-mm-ddThh-MM-ss.
    # Nos interesa solo la parte de fecha, y no la hora. Se hace un
    # split por la letra 'T' y nos quedamos con el primer elemento.
    date_string = date_string.split('T')[0]

    # Crea un objeto datetime a partir del formato especificado
    return datetime.strptime(date_string, "%Y-%m-%d")


def clean_str(s):
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        ":": "",
        ".": ""
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.lower().strip()


def validate_url(uri_string):
    """Valida si un string es una URI válida."""
    try:
        result = urlparse(uri_string)
        has_elements = all([result.scheme, result.netloc, result.path])
        is_http = result.scheme == "http" or result.scheme == "https"
        return True if has_elements and is_http else False

    except Exception:
        return False


def ensure_dir_exists(directory):
    """Se asegura de que un directorio exista."""
    if directory and not os.path.exists(directory):
        os.makedirs(directory)


def traverse_dict(dicc, keys, default_value=None):
    """Recorre un diccionario siguiendo una lista de claves, y devuelve
    default_value en caso de que alguna de ellas no exista.

    Args:
        dicc (dict): Diccionario a ser recorrido.
        keys (list): Lista de claves a ser recorrida. Puede contener
            índices de listas y claves de diccionarios mezcladas.
        default_value: Valor devuelto en caso de que `dicc` no se pueda
            recorrer siguiendo secuencialmente la lista de `keys` hasta
            el final.

    Returns:
        object: El valor obtenido siguiendo la lista de `keys` dentro de
        `dicc`.
    """
    for key in keys:
        if isinstance(dicc, dict) and key in dicc:
            dicc = dicc[key]
        elif (isinstance(dicc, list) and isinstance(key, int)
              and key < len(dicc)):
            dicc = dicc[key]
        else:
            return default_value

    return dicc


def is_list_of_matching_dicts(list_of_dicts, expected_keys=None):
    """Comprueba que una lista esté compuesta únicamente por diccionarios,
    que comparten exactamente las mismas claves.

    Args:
        list_of_dicts (list): Lista de diccionarios a comparar.
        expected_keys (set): Conjunto de las claves que cada diccionario debe
            tener. Si no se incluye, se asume que son las claves del primer
            diccionario de la lista.

    Returns:
        bool: True si todos los diccionarios comparten sus claves.
    """

    if isinstance(list_of_dicts, list) and len(list_of_dicts) == 0:
        return False

    is_not_list_msg = """
{} no es una lista. Debe ingresar una lista""".format(list_of_dicts)
    assert isinstance(list_of_dicts, list), is_not_list_msg

    not_all_dicts_msg = """
No todos los elementos de {} son diccionarios. Ingrese una lista compuesta
solo por diccionarios.""".format(list_of_dicts)
    assert all([isinstance(d, dict) for d in list_of_dicts]), not_all_dicts_msg

    # Si no se pasan expected_keys, se las toma del primer diccionario
    expected_keys = expected_keys or set(list_of_dicts[0].keys())

    elements = [set(d.keys()) == expected_keys for d in list_of_dicts]

    return all(elements)


def parse_value(cell):
    """Extrae el valor de una celda de Excel como texto."""
    value = cell.value

    # stripea espacios en strings
    if isinstance(value, string_types):
        value = value.strip()

    # convierte a texto ISO 8601 las fechas
    if isinstance(value, (datetime)):
        value = value.isoformat()

    return value


def sheet_to_table(worksheet):
    """Transforma una hoja de libro de Excel en una lista de diccionarios.

    Args:
        worksheet (Workbook.worksheet): Hoja de cálculo de un archivo XLSX
            según los lee `openpyxl`

    Returns:
        list_of_dicts: Lista de diccionarios, con tantos elementos como
            registros incluya la hoja, y con tantas claves por diccionario como
            campos tenga la hoja.
    """

    headers = []
    value_rows = []
    for row_i, row in enumerate(worksheet.iter_rows()):

        # lee los headers y el tamaño máximo de la hoja en columnas en fila 1
        if row_i == 0:
            for header_cell in row:
                if header_cell.value:
                    headers.append(parse_value(header_cell))
                else:
                    break
            continue

        # limita la cantidad de celdas a considerar, por la cantidad de headers
        row_cells = [
            parse_value(cell) for index, cell in enumerate(row)
            if index < len(headers)
        ]

        # agrega las filas siguientes que tengan al menos un campo no nulo
        if any(row_cells):
            value_rows.append(row_cells)
        # no se admiten filas vacías, eso determina el fin de la hoja
        else:
            break

    # convierte las filas en diccionarios con los headers como keys
    table = [
        # Ignoro los campos con valores nulos (None)
        {k: v
         for (k, v) in zip(headers, row) if v is not None}
        for row in value_rows
    ]

    return table


def string_to_list(string, sep=","):
    """Transforma una string con elementos separados por `sep` en una lista."""
    return [value.strip() for value in string.split(sep)]


def add_dicts(one_dict, other_dict):
    """Suma clave a clave los dos diccionarios. Si algún valor es un
    diccionario, llama recursivamente a la función. Ambos diccionarios deben
    tener exactamente las mismas claves, y los valores asociados deben ser
    sumables, o diccionarios.

    Args:
        one_dict (dict)
        other_dict (dict)

    Returns:
        dict: resultado de la suma
    """
    result = other_dict.copy()
    for k, v in list(one_dict.items()):
        if v is None:
            v = 0

        if isinstance(v, dict):
            result[k] = add_dicts(v, other_dict.get(k, {}))
        else:
            other_value = result.get(k, 0)
            if other_value is None:
                other_value = 0
            result[k] = other_value + v

    return result


def parse_repeating_time_interval(date_str, to="days"):
    if to == "days":
        return parse_repeating_time_interval_to_days(date_str)

    elif to == "string":
        return parse_repeating_time_interval_to_str(date_str)

    else:
        raise NotImplementedError("Falta implementar parsing a {}".format(to))


def parse_repeating_time_interval_to_days(date_str):
    """Parsea un string con un intervalo de tiempo con repetición especificado
    por la norma ISO 8601 en una cantidad de días que representa ese intervalo.
    Devuelve 0 en caso de que el intervalo sea inválido.
    """

    intervals = {'Y': 365, 'M': 30, 'W': 7, 'D': 1, 'H': 0, 'S': 0}

    if date_str.find('R/P') != 0:  # Periodicity mal formada
        return 0

    date_str = date_str.strip('R/P')
    days = 0
    index = 0
    for interval in intervals:
        value_end = date_str.find(interval)
        if value_end < 0:
            continue
        try:
            days += int(float(date_str[index:value_end]) * intervals[interval])
        # Valor de accrualPeriodicity inválido, se toma como 0
        except ValueError:
            continue
        index = value_end

    # Si el número de días es menor lo redondeamos a 1
    return max(days, 1)


def parse_repeating_time_interval_to_str(date_str):
    """Devuelve descripción humana de un intervalo de repetición.

    TODO: Por ahora sólo interpreta una lista fija de intervalos. Debería poder
    parsear cualquier caso.
    """

    with open(
            os.path.join(ABSOLUTE_SCHEMA_DIR, "accrualPeriodicity.json"),
            "r") as f:
        freqs_map = {freq["id"]: freq["description"] for freq in json.load(f)}

    return freqs_map[date_str]


def get_ws_case_insensitive(wb, title):
    """Devuelve una hoja en un workbook sin importar mayúsculas/minúsculas."""
    return wb[find_ws_name(wb, title)]


def find_ws_name(wb, name):
    """Busca una hoja en un workbook sin importar mayúsculas/minúsculas."""
    if isinstance(wb, string_types):
        # FIXME: importar o borrar segun corresponda
        wb = load_workbook(wb, read_only=True, data_only=True)

    for sheetname in wb.sheetnames:
        if sheetname.lower() == name.lower():
            return sheetname

    raise Exception("No existe la hoja {}".format(name))
